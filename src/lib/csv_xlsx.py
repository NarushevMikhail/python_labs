import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:

    csv_file = Path(csv_path)
    xlsx_file = Path(xlsx_path)
    
    if not csv_file.exists():
        raise FileNotFoundError(f"Файл {csv_path} не найден")
    
    if csv_file.suffix.lower() != '.csv':
        raise ValueError("Неверный тип файла. Ожидается .csv")
    
    # Чтение CSV
    try:
        with csv_file.open('r', encoding='utf-8') as f:
            reader = csv.reader(f) #создает объект для чтения CSV
            rows = list(reader) #преобразует все строки CSV в список списков
    except Exception as e:
        raise ValueError(f"Ошибка чтения CSV: {e}")
    
    if not rows: #проверяет что файл не пустой
        raise ValueError("Пустой CSV файл")
    
    if not rows[0]:
        raise ValueError("CSV файл не содержит заголовка")
    
    try: # Создание XLSX
        wb = Workbook() #создает новую Excel книгу
        ws = wb.active #получает активный лист
        ws.title = "Sheet1" #название листа
        
        for row in rows:
            ws.append(row) #добавляет целую строку данных в лист
        
        
        for col_idx, column_cells in enumerate(ws.columns, 1):  #перебирает столбцы листа, начиная с индекса 1
            max_length = 8  
            column_letter = get_column_letter(col_idx) #преобразует номер столбца в букву
            
            for cell in column_cells:
                try:
                    if cell.value: #значение ячейки

                        cell_length = len(str(cell.value)) #длину содержимого ячейки
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = max_length + 2 #добавляет 2 символа для отступов
            ws.column_dimensions[column_letter].width = adjusted_width #устанавливает ширину столбца
        
        wb.save(xlsx_file)
        
    except Exception as e:
        raise ValueError(f"Ошибка создания XLSX: {e}")
    
# csv_to_xlsx("data/samples/people.csv", "data/out/people.xlsx")
